import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict, Any
import io
import logging
from datetime import datetime
import json

logger = logging.getLogger(__name__)

class ExcelService:
    """Service for creating Excel files with structured content"""
    
    def __init__(self):
        self.headers = ['Topic', 'Subtopic', 'Content', 'Video Link']
    
    def create_excel_file(self, data: List[Dict[str, Any]], filename: str = None, detailed_content: Dict[str, Any] = None) -> bytes:
        """
        Create an Excel file with the structured content
        """
        try:
            # Parse and clean the data using the new parsing function
            cleaned_data = self.parse_json_data(data)

            # Add formulas as rows in the main content sheet with MathJax formatting
            if detailed_content and detailed_content.get('formulas'):
                logger.info(f"Processing {len(detailed_content['formulas'])} formulas for MathJax conversion")
                for i, formula in enumerate(detailed_content['formulas']):
                    logger.info(f"Formula {i+1}: {formula}")
                    latex_content = formula.get('latex', '')
                    if latex_content:
                        # Format as MathJax for web rendering
                        mathjax_content = f"$${latex_content}$$"
                        logger.info(f"Converted to MathJax: {mathjax_content}")
                    else:
                        mathjax_content = formula.get('description', 'Formula')
                        logger.info(f"No LaTeX found, using description: {mathjax_content}")
                    
                    cleaned_data.append({
                        'Topic': formula.get('topic', 'Formula'),
                        'Subtopic': formula.get('description', 'Formula'),
                        'Content': mathjax_content,
                        'Video Link': ''
                    })
            else:
                logger.info("No formulas found in detailed_content")
                if detailed_content:
                    logger.info(f"Available content types: {list(detailed_content.keys())}")

            # Create DataFrame with cleaned data
            df = pd.DataFrame(cleaned_data)

            # Ensure we have the required columns in the right order
            required_columns = ['Topic', 'Subtopic', 'Content', 'Video Link']
            for col in required_columns:
                if col not in df.columns:
                    df[col] = ''
            df = df[required_columns]

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Content', index=False)
                workbook = writer.book
                worksheet = writer.sheets['Content']
                self._apply_formatting(worksheet, df)
                if detailed_content:
                    self._add_detailed_content_sheets(workbook, detailed_content)
            output.seek(0)
            return output.getvalue()
        except Exception as e:
            logger.error(f"Error creating Excel file: {str(e)}")
            raise
    
    def _add_detailed_content_sheets(self, workbook, detailed_content: Dict[str, Any]):
        """Add detailed content sheets for tables, formulas, and diagrams"""
        try:
            # Tables sheet
            if detailed_content.get('tables'):
                self._add_tables_sheet(workbook, detailed_content['tables'])
            
            # Formulas sheet
            if detailed_content.get('formulas'):
                self._add_formulas_sheet(workbook, detailed_content['formulas'])
            
            # Diagrams sheet
            if detailed_content.get('diagrams'):
                self._add_diagrams_sheet(workbook, detailed_content['diagrams'])
                
        except Exception as e:
            logger.error(f"Error adding detailed content sheets: {str(e)}")
    
    def _add_tables_sheet(self, workbook, tables: List[Dict[str, Any]]):
        """Add tables to a separate sheet"""
        try:
            # Create tables data
            table_data = []
            for i, table in enumerate(tables):
                if 'table_data' in table and table['table_data'].get('rows'):
                    for row in table['table_data']['rows']:
                        row_data = {
                            'Table_Number': i + 1,
                            'Page': table.get('page', 'Unknown'),
                            **row
                        }
                        table_data.append(row_data)
            
            if table_data:
                df_tables = pd.DataFrame(table_data)
                df_tables.to_excel(workbook, sheet_name='Tables', index=False)
                
                # Format tables sheet
                worksheet = workbook['Tables']
                self._apply_formatting(worksheet, df_tables)
                
        except Exception as e:
            logger.error(f"Error adding tables sheet: {str(e)}")
    
    def _add_formulas_sheet(self, workbook, formulas: List[Dict[str, Any]]):
        """Add formulas to a separate sheet"""
        try:
            formula_data = []
            for i, formula in enumerate(formulas):
                formula_data.append({
                    'Formula_Number': i + 1,
                    'Page': formula.get('page', 'Unknown'),
                    'LaTeX': formula.get('latex', ''),
                    'Description': formula.get('description', ''),
                    'Variables': formula.get('variables', '')
                })
            
            if formula_data:
                df_formulas = pd.DataFrame(formula_data)
                df_formulas.to_excel(workbook, sheet_name='Formulas', index=False)
                
                # Format formulas sheet
                worksheet = workbook['Formulas']
                self._apply_formatting(worksheet, df_formulas)
                
        except Exception as e:
            logger.error(f"Error adding formulas sheet: {str(e)}")
    
    def _add_diagrams_sheet(self, workbook, diagrams: List[Dict[str, Any]]):
        """Add diagrams to a separate sheet"""
        try:
            diagram_data = []
            for i, diagram in enumerate(diagrams):
                diagram_data.append({
                    'Diagram_Number': i + 1,
                    'Page': diagram.get('page', 'Unknown'),
                    'Description': diagram.get('description', ''),
                    'Content_Type': diagram.get('content_type', 'diagram')
                })
            
            if diagram_data:
                df_diagrams = pd.DataFrame(diagram_data)
                df_diagrams.to_excel(workbook, sheet_name='Diagrams', index=False)
                
                # Format diagrams sheet
                worksheet = workbook['Diagrams']
                self._apply_formatting(worksheet, df_diagrams)
                
        except Exception as e:
            logger.error(f"Error adding diagrams sheet: {str(e)}")
    
    def _apply_formatting(self, worksheet, df):
        """Apply formatting to the Excel worksheet"""
        try:
            # Define styles
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            content_font = Font(size=11)
            content_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format headers
            for col in range(1, len(self.headers) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Format content cells
            for row in range(2, len(df) + 2):
                for col in range(1, len(self.headers) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = content_font
                    cell.alignment = content_alignment
                    cell.border = border
            
            # Set column widths
            column_widths = {
                'A': 30,  # Topic
                'B': 25,  # Subtopic
                'C': 60,  # Content
                'D': 20   # Video Link
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Set row heights for better readability
            for row in range(2, len(df) + 2):
                worksheet.row_dimensions[row].height = 60
            
            # Freeze the header row
            worksheet.freeze_panes = "A2"
            
        except Exception as e:
            logger.error(f"Error applying formatting: {str(e)}")
    
    def create_advanced_excel(self, data: List[Dict[str, Any]], metadata: Dict[str, Any] = None, detailed_content: Dict[str, Any] = None) -> bytes:
        """
        Create an advanced Excel file with multiple sheets and metadata
        """
        try:
            output = io.BytesIO()
            cleaned_data = self.parse_json_data(data)
            # Add formulas as rows in the main content sheet with MathJax formatting
            if detailed_content and detailed_content.get('formulas'):
                logger.info(f"Advanced Excel: Processing {len(detailed_content['formulas'])} formulas for MathJax conversion")
                for i, formula in enumerate(detailed_content['formulas']):
                    logger.info(f"Advanced Excel: Formula {i+1}: {formula}")
                    latex_content = formula.get('latex', '')
                    if latex_content:
                        # Format as MathJax for web rendering
                        mathjax_content = f"$${latex_content}$$"
                        logger.info(f"Advanced Excel: Converted to MathJax: {mathjax_content}")
                    else:
                        mathjax_content = formula.get('description', 'Formula')
                        logger.info(f"Advanced Excel: No LaTeX found, using description: {mathjax_content}")
                    
                    cleaned_data.append({
                        'Topic': formula.get('topic', 'Formula'),
                        'Subtopic': formula.get('description', 'Formula'),
                        'Content': mathjax_content,
                        'Video Link': ''
                    })
            else:
                logger.info("Advanced Excel: No formulas found in detailed_content")
                if detailed_content:
                    logger.info(f"Advanced Excel: Available content types: {list(detailed_content.keys())}")
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_content = pd.DataFrame(cleaned_data)
                required_columns = ['Topic', 'Subtopic', 'Content', 'Video Link']
                for col in required_columns:
                    if col not in df_content.columns:
                        df_content[col] = ''
                df_content = df_content[required_columns]
                df_content.to_excel(writer, sheet_name='Content', index=False)
                if metadata:
                    df_metadata = pd.DataFrame([metadata])
                    df_metadata.to_excel(writer, sheet_name='Metadata', index=False)
                summary_data = {
                    'Metric': [
                        'Total Topics', 'Total Subtopics', 'Total Content Items', 'Tables Found',
                        'Formulas Found', 'Diagrams Found', 'Generated Date'
                    ],
                    'Value': [
                        len([item for item in cleaned_data if item.get('Topic')]),
                        len([item for item in cleaned_data if item.get('Subtopic')]),
                        len(cleaned_data),
                        metadata.get('content_breakdown', {}).get('tables', 0) if metadata else 0,
                        metadata.get('content_breakdown', {}).get('formulas', 0) if metadata else 0,
                        metadata.get('content_breakdown', {}).get('diagrams', 0) if metadata else 0,
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    ]
                }
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
                workbook = writer.book
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    if sheet_name == 'Content':
                        df = df_content
                    elif sheet_name == 'Metadata':
                        df = df_metadata
                    elif sheet_name == 'Summary':
                        df = df_summary
                    else:
                        continue
                    self._apply_formatting(worksheet, df)
                if detailed_content:
                    self._add_detailed_content_sheets(workbook, detailed_content)
            output.seek(0)
            return output.getvalue()
        except Exception as e:
            logger.error(f"Error creating advanced Excel file: {str(e)}")
            raise
    
    def parse_json_data(self, data: List[Any]) -> List[Dict[str, str]]:
        """
        Parse JSON data and convert to proper Excel format
        
        Args:
            data: List of data items (could be dicts, strings, or JSON strings)
            
        Returns:
            List of properly formatted dictionaries for Excel
        """
        parsed_data = []
        
        for item in data:
            # If item is a string that looks like JSON, try to parse it
            if isinstance(item, str):
                try:
                    # Clean the string
                    cleaned_item = item.strip()
                    if cleaned_item.startswith('```json'):
                        cleaned_item = cleaned_item[7:]
                    if cleaned_item.endswith('```'):
                        cleaned_item = cleaned_item[:-3]
                    cleaned_item = cleaned_item.strip()
                    
                    # Try to parse as JSON
                    parsed = json.loads(cleaned_item)
                    if isinstance(parsed, list):
                        # If it's a list, process each item
                        for list_item in parsed:
                            if isinstance(list_item, dict):
                                parsed_data.append({
                                    'Topic': str(list_item.get('topic', '')).strip(),
                                    'Subtopic': str(list_item.get('subtopic', '')).strip(),
                                    'Content': str(list_item.get('content', '')).strip(),
                                    'Video Link': str(list_item.get('video_link', '')).strip()
                                })
                    elif isinstance(parsed, dict):
                        parsed_data.append({
                            'Topic': str(parsed.get('topic', '')).strip(),
                            'Subtopic': str(parsed.get('subtopic', '')).strip(),
                            'Content': str(parsed.get('content', '')).strip(),
                            'Video Link': str(parsed.get('video_link', '')).strip()
                        })
                    else:
                        # If parsing fails, treat as content
                        parsed_data.append({
                            'Topic': '',
                            'Subtopic': '',
                            'Content': str(item).strip(),
                            'Video Link': ''
                        })
                except:
                    # If JSON parsing fails, treat as content
                    parsed_data.append({
                        'Topic': '',
                        'Subtopic': '',
                        'Content': str(item).strip(),
                        'Video Link': ''
                    })
            elif isinstance(item, dict):
                # If item is already a dict, use it directly
                parsed_data.append({
                    'Topic': str(item.get('topic', item.get('Topic', ''))).strip(),
                    'Subtopic': str(item.get('subtopic', item.get('Subtopic', ''))).strip(),
                    'Content': str(item.get('content', item.get('Content', ''))).strip(),
                    'Video Link': str(item.get('video_link', item.get('Video Link', item.get('Video_Link', '')))).strip()
                })
            else:
                # Fallback for any other type
                parsed_data.append({
                    'Topic': '',
                    'Subtopic': '',
                    'Content': str(item).strip(),
                    'Video Link': ''
                })
        
        return parsed_data

    def validate_data(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Validate the data structure before creating Excel
        
        Args:
            data: List of dictionaries to validate
            
        Returns:
            Validation result
        """
        validation_result = {
            'valid': True,
            'errors': [],
            'warnings': [],
            'stats': {
                'total_items': len(data),
                'items_with_topic': 0,
                'items_with_subtopic': 0,
                'items_with_content': 0,
                'items_with_video_link': 0
            }
        }
        
        for i, item in enumerate(data):
            # Check required fields
            if not item.get('Topic'):
                validation_result['warnings'].append(f"Item {i+1}: Missing Topic")
            else:
                validation_result['stats']['items_with_topic'] += 1
            
            if not item.get('Subtopic'):
                validation_result['warnings'].append(f"Item {i+1}: Missing Subtopic")
            else:
                validation_result['stats']['items_with_subtopic'] += 1
            
            if not item.get('Content'):
                validation_result['errors'].append(f"Item {i+1}: Missing Content")
            else:
                validation_result['stats']['items_with_content'] += 1
            
            if item.get('Video Link'):
                validation_result['stats']['items_with_video_link'] += 1
        
        if validation_result['errors']:
            validation_result['valid'] = False
        
        return validation_result 